from dash import Dash, html, dcc
import pandas as pd
import plotly.express as px
import dash_mantine_components as dmc
import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timedelta
import random

# Manipulação de dados DESCONTO/VENDAS
arquivo_excel = "Controle.xlsx"
df = pd.read_excel(arquivo_excel)

df["Preço de Venda Atacado"] = df["Preço de Venda Atacado"].apply(lambda x: float(str(x).replace("R$", "").replace(" ", "").replace(",",".")))
df["Preço de Venda"] = df["Preço de Venda"].apply(lambda x: float(str(x).replace("R$", "").replace(" ", "").replace(",",".")))


def vendas():
    df["Vendas"].sort_values(ascending=False)
    novo_arquivo_excel = "vendas.xlsx"
    dados_ordenados = df.sort_values(by="Vendas", ascending=False)
    dados_ordenados.to_excel(novo_arquivo_excel, index=False)

def desconto_xlsx():

 df["Desconto"] = df["Preço de Venda"] - df["Preço de Venda Atacado"]
 novo_arquivo_excel = "desconto.xlsx"
 dados_ordenados = df.sort_values(by="Desconto", ascending=False)
 dados_ordenados.to_excel(novo_arquivo_excel, index=False)
 


desconto_xlsx()
vendas()

a = "desconto.xlsx"
b = "vendas.xlsx"

top_20_produtos = pd.read_excel(a).head(20)
top_10_vendas = pd.read_excel(b).head(10)

# Manipulação de dados GRÁFICO DE SIMULAÇÃO

arquivo_origem = "Controle.xlsx"
workbook_origem = openpyxl.load_workbook(arquivo_origem)
planilha_origem = workbook_origem.active

workbook_destino = openpyxl.Workbook()
planilha_destino = workbook_destino.active
planilha_destino.title = "Dados Associados"

planilha_destino['A1'] = "Data"
planilha_destino['B1'] = "Descrição"
planilha_destino['A1'].font = Font(bold=True)
planilha_destino['B1'].font = Font(bold=True)

data_inicial = datetime(2023, 11, 1)
data_final = datetime(2023, 11, 30)

numero_de_linhas = 5000

descricoes_origem = [cell.value for cell in planilha_origem['B'][1:planilha_origem.max_row]]

planilha_destino.column_dimensions['A'].width = 12 

for linha in range(2, numero_de_linhas + 2):
    data_aleatoria = data_inicial + timedelta(days=random.randint(0, (data_final - data_inicial).days))
    descricao_aleatoria = random.choice(descricoes_origem)
    
    planilha_destino.cell(row=linha, column=1, value=data_aleatoria)
    planilha_destino.cell(row=linha, column=2, value=descricao_aleatoria)

arquivo_destino = "dados_associados.xlsx"
workbook_destino.save(arquivo_destino)
print(f"Arquivo Excel '{arquivo_destino}' criado com 5000 linhas de dados associados às descrições da planilha de origem.")

arquivo_excel_simulação = "dados_associados.xlsx"
dfs = pd.read_excel(arquivo_excel_simulação)


dfs["Data"] = dfs["Data"].apply(lambda x: str(x).replace("00:00:00", "").replace(" ", ""))


vendas_por_data = dfs.groupby('Data').size().reset_index(name='Total_Vendas')

# Frontend
external_stylesheets = [dmc.theme.DEFAULT_COLORS]
app = Dash(__name__, external_stylesheets=external_stylesheets)

style = {
    'height': '100',
    'background': '#f6f8fc'
}

app.layout = dmc.Container([
    dmc.Title('Gráficos de controle', size='h4', style={'textAlign': 'center', 'padding': '15px'}),

    dmc.Grid([
        dmc.Col([
            dmc.Card(
                children=[
                    dmc.Group(
                        [
                            dmc.Text("Top 20 produtos com maiores descontos no atacado", weight=500),
                            dmc.Badge("GRÁFICO EM AREA", color="blue", variant="light"),
                        ],
                        position="apart",
                        mt="md",
                        mb="xs",
                    ),
                    dmc.CardSection(
                        dcc.Graph(figure=px.area(top_20_produtos, x='Produtos', y='Desconto'))
                    ),
                ],
                withBorder=True,
                shadow="sm",
                radius="md"
            )
        ], span=12),

        dmc.Col([
            dmc.Card(
                children=[
                    dmc.Group(
                        [
                            dmc.Text("Top 10 produtos com maiores vendas", weight=500),
                            dmc.Badge("GRÁFICO EM BARRA", color="blue", variant="light"),
                        ],
                        position="apart",
                        mt="md",
                        mb="xs",
                    ),
                    dmc.CardSection(
                        dcc.Graph(figure=px.bar(top_10_vendas,  x='Produtos', y='Vendas'))
                    )
                ],
                withBorder=True,
                shadow="sm",
                radius="md"
            )
        ], span=12),

        dmc.Col([
            dmc.Card(
                children=[
                    dmc.Group(
                        [
                            dmc.Text("Total de vendas por data", weight=500),
                            dmc.Badge("GRÁFICO EM LINHA", color="blue", variant="light"),
                        ],
                        position="apart",
                        mt="md",
                        mb="xs",
                    ),
                    dmc.CardSection(
                        dcc.Graph(figure=px.line(vendas_por_data,  x='Data', y='Total_Vendas'))
                    )
                ],
                withBorder=True,
                shadow="sm",
                radius="md"
            )
        ], span=12)
    ])
], fluid=True, style=style )

if __name__ == '__main__':
    app.run(debug=True)
